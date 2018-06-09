from openpyxl import load_workbook

filepath = '/home/iroan/data/carnum.xlsx'

# 读取sheet名称
# print(wb2.sheetnames)

def task_0609(filename):
    '''
    功能：提取出每一行数据中A字段对应的省份在C中出现的字母集合
    :param filename: 处理的文件名
    :return: 一个json格式的数据
        数据形式：
                [
                    {"省份X":'字幕X','...'}
                    ...
                ]

    实现步骤:
        1. 获取到省份列，字母列数据
        1. 获取省份的集合province_set
        1. 根据集合province_set的字段创建字典
        3. 判断字母是否在 province_set中，如果在，添加到集合中
    '''
    wb = load_workbook(filepath)
    sheet = wb['Sheet1']
    province_set = set()
    for index,i in enumerate(sheet['A']):
        if index == 1:
            continue
        if i.value != None:
            province_set.add(i.value)

    result_dict = dict()


task_0609(filepath)
